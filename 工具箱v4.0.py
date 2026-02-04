# -*- coding: utf-8 -*-
# ==================== 导入依赖库 ====================
import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from threading import Thread
from queue import Queue
from PIL import Image  # 图像处理库
import openpyxl  # Excel处理库
from openpyxl import load_workbook
from openpyxl.styles import Font
import re
import json  # 用于保存配置文件

# ==================== 全局样式配置 ====================
COLORS = {
    'background': '#F0F0F0',  # 背景色
    'primary': '#4A90E2',  # 主色调（蓝色）
    'secondary': '#50E3C2',  # 辅助色（青色）
    'warning': '#F5A623',  # 警告色（橙色）
    'danger': '#D0021B',  # 危险色（红色）
    'text': '#333333',  # 文字颜色
    'success': '#7ED321'  # 成功色（绿色）
}


# ==================== 主应用程序类 ====================
class MainApplication:
    def __init__(self, master):
        """初始化主窗口"""
        self.master = master
        master.title("多功能文件处理工具箱 v4.0")
        master.geometry("1000x800")
        self.configure_styles()

        # 创建选项卡容器
        self.notebook = ttk.Notebook(master)
        self.notebook.pack(expand=1, fill="both", padx=10, pady=10)

        # 初始化各功能模块
        self.rename_module = RenameModule(self.notebook)  # 文件重命名模块
        self.convert_module = ConvertModule(self.notebook)  # 图片格式转换模块
        self.hyperlink_module = HyperlinkModule(self.notebook)  # 超链接转换模块

        # 添加选项卡标签
        self.notebook.add(self.rename_module.frame, text="📁 批量重命名")
        self.notebook.add(self.convert_module.frame, text="🖼️ 图片格式转换")
        self.notebook.add(self.hyperlink_module.frame, text="🔗 Excel超链接转换")

    def configure_styles(self):
        """配置全局控件样式"""
        style = ttk.Style()
        style.theme_use('clam')  # 使用clam主题

        # 配置选项卡样式
        style.configure('TNotebook', background=COLORS['background'])
        style.configure('TNotebook.Tab',
                        font=('微软雅黑', 10, 'bold'),
                        padding=[15, 5],
                        background=COLORS['secondary'],
                        foreground=COLORS['text'])
        style.map('TNotebook.Tab',
                  background=[('selected', COLORS['primary'])],
                  foreground=[('selected', 'white')])

        # 配置按钮样式
        style.configure('Primary.TButton',
                        font=('微软雅黑', 9),
                        background=COLORS['primary'],
                        foreground='white',
                        bordercolor=COLORS['primary'],
                        focuscolor=COLORS['primary'])
        style.map('Primary.TButton',
                  background=[('active', '#357ABD')],
                  foreground=[('active', 'white')])


# ==================== 模块基类 ====================
class BaseModule:
    """所有功能模块的基类"""

    def __init__(self, parent):
        self.frame = ttk.Frame(parent)  # 模块主框架
        self.log_queue = Queue()  # 日志消息队列
        self.running = False  # 任务运行状态
        self.create_widgets()  # 创建界面组件
        self.process_log_queue()  # 启动日志处理

    def process_log_queue(self):
        """实时处理日志队列"""
        while not self.log_queue.empty():
            msg_type, content = self.log_queue.get()
            if msg_type == "end":
                break
            self.log_area.config(state=tk.NORMAL)
            self.log_area.insert(tk.END, content + "\n", msg_type)
            self.log_area.see(tk.END)  # 自动滚动到底部
            self.log_area.config(state=tk.DISABLED)
        self.frame.after(100, self.process_log_queue)

    def clear_log(self):
        """清空日志内容"""
        self.log_area.config(state=tk.NORMAL)
        self.log_area.delete(1.0, tk.END)
        self.log_area.config(state=tk.DISABLED)


# ==================== 文件重命名模块 ====================
class RenameModule(BaseModule):
    """批量文件重命名功能（支持修改后缀）"""

    def __init__(self, parent):
        super().__init__(parent)

    def create_widgets(self):
        """构建界面组件"""
        # ----- 使用说明 -----
        help_text = """使用说明：
1. 选择需要批量重命名的文件夹
2. 设置文件名前缀、排序方式、序号位数
3. 【新增】可设置统一文件后缀（如：.txt）
4. 点击【开始重命名】执行操作"""
        ttk.Label(self.frame, text=help_text, foreground=COLORS['text']).pack(pady=5, anchor="w")

        # ----- 目录选择部分 -----
        dir_frame = ttk.Frame(self.frame)
        dir_frame.pack(pady=10, fill=tk.X, padx=15)

        ttk.Label(dir_frame, text="目标目录:", font=('微软雅黑', 9)).pack(side=tk.LEFT)
        self.dir_entry = ttk.Entry(dir_frame, width=50)  # 目录输入框
        self.dir_entry.pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)
        ttk.Button(dir_frame, text="浏览...", style='Primary.TButton', command=self.browse_dir).pack(side=tk.LEFT)

        # ----- 参数设置部分 -----
        param_frame = ttk.LabelFrame(self.frame, text="重命名参数", style='Card.TLabelframe')
        param_frame.pack(pady=10, fill=tk.X, padx=15)

        # 前缀设置
        ttk.Label(param_frame, text="前缀:").grid(row=0, column=0, padx=5, pady=3)
        self.prefix_entry = ttk.Entry(param_frame)
        self.prefix_entry.insert(0, "file")  # 默认前缀
        self.prefix_entry.grid(row=0, column=1, sticky=tk.W, padx=5)

        # 排序方式
        ttk.Label(param_frame, text="排序方式:").grid(row=0, column=2, padx=5)
        self.sort_combo = ttk.Combobox(param_frame,
                                       values=["名称", "修改时间", "创建时间"],
                                       state="readonly")
        self.sort_combo.current(0)  # 默认选择第一个
        self.sort_combo.grid(row=0, column=3, padx=5)

        # 序号位数
        ttk.Label(param_frame, text="序号位数:").grid(row=0, column=4, padx=5)
        self.digits_spin = ttk.Spinbox(param_frame,
                                       from_=1,  # 最小值
                                       to=6,  # 最大值
                                       width=5)
        self.digits_spin.set(3)  # 默认3位
        self.digits_spin.grid(row=0, column=5, padx=5)

        # 新增后缀设置
        ttk.Label(param_frame, text="后缀:").grid(row=0, column=6, padx=5)
        self.suffix_entry = ttk.Entry(param_frame, width=8)
        self.suffix_entry.grid(row=0, column=7, padx=5)

        # ----- 操作按钮 -----
        btn_frame = ttk.Frame(self.frame)
        btn_frame.pack(pady=10)
        self.start_btn = ttk.Button(btn_frame,
                                    text="▶ 开始重命名",
                                    style='Primary.TButton',
                                    command=self.start_rename)
        self.start_btn.pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="🗑️ 清空日志", command=self.clear_log).pack(side=tk.LEFT)

        # ----- 日志区域 -----
        log_frame = ttk.LabelFrame(self.frame, text="操作日志", style='Card.TLabelframe')
        log_frame.pack(pady=10, fill=tk.BOTH, expand=True, padx=15)

        self.log_area = scrolledtext.ScrolledText(log_frame,
                                                  height=15,
                                                  wrap=tk.WORD)  # 自动换行
        self.log_area.pack(fill=tk.BOTH, expand=True)

        # 配置日志标签颜色
        self.log_area.tag_config("success", foreground=COLORS['success'])
        self.log_area.tag_config("warning", foreground=COLORS['warning'])
        self.log_area.tag_config("error", foreground=COLORS['danger'])
        self.log_area.config(state=tk.DISABLED)  # 初始禁用编辑

    def browse_dir(self):
        """选择目标目录"""
        directory = filedialog.askdirectory()
        if directory:
            self.dir_entry.delete(0, tk.END)
            self.dir_entry.insert(0, directory)

    def start_rename(self):
        """启动重命名任务"""
        if self.running:
            messagebox.showwarning("操作进行中", "当前已有任务正在运行，请稍候")
            return

        directory = self.dir_entry.get()
        if not os.path.isdir(directory):
            messagebox.showerror("错误", "无效的目录路径")
            return

        self.running = True
        self.start_btn.config(text="⏳ 运行中...", state=tk.DISABLED)

        # 收集参数
        params = {
            'directory': directory,
            'prefix': self.prefix_entry.get(),
            'sort_by': ['name', 'modified', 'created'][self.sort_combo.current()],
            'padding': int(self.digits_spin.get()),
            'suffix': self.suffix_entry.get().strip()  # 新增后缀参数
        }

        # 启动后台线程
        Thread(target=self.batch_rename, kwargs=params, daemon=True).start()

    def batch_rename(self, directory, prefix='item', sort_by='name', padding=3, suffix=''):
        """执行批量重命名核心逻辑（新增后缀处理）"""
        try:
            # 获取目录下所有条目（排除.和..）
            items = [item for item in os.listdir(directory)
                     if item not in ('.', '..')]

            # 定义排序方式对应的键函数
            sort_keys = {
                'name': lambda x: x.lower(),  # 按名称排序（不区分大小写）
                'modified': lambda x: os.path.getmtime(os.path.join(directory, x)),
                'created': lambda x: os.path.getctime(os.path.join(directory, x))
            }

            try:
                items_sorted = sorted(items, key=sort_keys[sort_by])
            except KeyError:
                self.log_queue.put(("error", f"无效的排序方式：'{sort_by}'，使用默认名称排序"))
                items_sorted = sorted(items, key=sort_keys['name'])

            # 倒序处理避免覆盖问题
            for idx in reversed(range(len(items_sorted))):
                old_name = items_sorted[idx]
                old_path = os.path.join(directory, old_name)

                # 处理后缀逻辑
                if suffix:  # 如果用户输入了后缀
                    if not suffix.startswith('.'):  # 自动补全点号
                        suffix = '.' + suffix
                    ext = suffix
                else:  # 保留原后缀
                    ext = os.path.splitext(old_name)[1] if os.path.isfile(old_path) else ''

                # ==== 修正后的关键代码 ====
                new_name = f"{prefix}_{idx + 1:0{padding}d}{ext}"
                # ========================

                new_path = os.path.join(directory, new_name)

                if old_path == new_path:
                    continue  # 无需重命名

                if os.path.exists(new_path):
                    self.log_queue.put(("warning", f"冲突：'{new_name}' 已存在，跳过"))
                    continue

                try:
                    os.rename(old_path, new_path)
                    self.log_queue.put(("success", f"{old_name} → {new_name}"))
                except Exception as e:
                    self.log_queue.put(("error", f"处理 {old_name} 失败 - {str(e)}"))

        except Exception as e:
            self.log_queue.put(("error", f"发生未预期错误：{str(e)}"))
        finally:
            self.log_queue.put(("end", ""))  # 结束标志
            self.running = False
            # 恢复按钮状态
            self.frame.after(100, lambda: self.start_btn.config(
                text="▶ 开始重命名",
                state=tk.NORMAL
            ))


# ==================== 图片格式转换模块 ====================
class ConvertModule(BaseModule):
    """图片格式批量转换功能"""

    def __init__(self, parent):
        self.input_files = []  # 待转换文件列表
        self.output_dir = ""  # 输出目录路径
        self.conversion_running = False  # 转换状态
        super().__init__(parent)

    def create_widgets(self):
        """构建界面组件"""
        # ----- 使用说明 -----
        help_text = """使用说明：
1. 选择图片文件或整个文件夹
2. 设置输出格式和目录
3. 点击【开始转换】执行操作"""
        ttk.Label(self.frame, text=help_text, foreground=COLORS['text']).pack(pady=5, anchor="w")

        # ----- 文件选择部分 -----
        file_frame = ttk.Frame(self.frame)
        file_frame.pack(pady=10, fill='x', padx=15)

        # 文件选择按钮
        ttk.Button(file_frame,
                   text="📄 选择文件",
                   style='Primary.TButton',
                   command=self.select_files).pack(side='left', padx=5)

        # 文件夹选择按钮
        ttk.Button(file_frame,
                   text="📁 选择文件夹",
                   style='Primary.TButton',
                   command=self.select_folder).pack(side='left', padx=5)

        # ----- 输出设置部分 -----
        settings_frame = ttk.LabelFrame(self.frame,
                                        text="输出设置",
                                        style='Card.TLabelframe')
        settings_frame.pack(pady=10, fill='x', padx=15)

        # 输出格式选择
        ttk.Label(settings_frame, text="输出格式:").pack(side='left', padx=5)
        self.output_formats = ['PNG', 'JPEG', 'BMP', 'WEBP', 'ICO']
        self.format_var = tk.StringVar(value='PNG')  # 默认格式
        format_combobox = ttk.Combobox(settings_frame,
                                       textvariable=self.format_var,
                                       values=self.output_formats,
                                       state="readonly",
                                       width=8)
        format_combobox.pack(side='left', padx=5)
        format_combobox.bind('<<ComboboxSelected>>', self.toggle_ico_settings)

        # ICO尺寸设置（默认隐藏）
        self.ico_frame = ttk.Frame(settings_frame)
        ttk.Label(self.ico_frame, text="尺寸:").pack(side='left')
        self.ico_sizes = ['16x16', '32x32', '48x48', '64x64', '128x128', '256x256']
        self.size_var = tk.StringVar(value='256x256')  # 默认尺寸
        ico_combobox = ttk.Combobox(self.ico_frame,
                                    textvariable=self.size_var,
                                    values=self.ico_sizes,
                                    state="readonly",
                                    width=8)
        ico_combobox.pack(side='left', padx=5)

        # ----- 输出目录选择 -----
        output_frame = ttk.Frame(self.frame)
        output_frame.pack(pady=10, fill='x', padx=15)
        ttk.Button(output_frame,
                   text="📂 输出目录",
                   style='Primary.TButton',
                   command=self.select_output_dir).pack(side='left', padx=5)
        self.output_label = ttk.Label(output_frame,
                                      text="未选择",
                                      foreground=COLORS['text'])
        self.output_label.pack(side='left', padx=5)

        # ----- 转换按钮 -----
        self.convert_btn = ttk.Button(self.frame,
                                      text="▶ 开始转换",
                                      style='Primary.TButton',
                                      command=self.start_conversion)
        self.convert_btn.pack(pady=10)

        # ----- 日志区域 -----
        log_frame = ttk.LabelFrame(self.frame,
                                   text="转换日志",
                                   style='Card.TLabelframe')
        log_frame.pack(pady=10, fill='both', expand=True, padx=15)
        self.log_area = scrolledtext.ScrolledText(log_frame,
                                                  wrap=tk.WORD,
                                                  height=12)
        self.log_area.pack(fill='both', expand=True)
        self.log_area.tag_config("success", foreground=COLORS['success'])
        self.log_area.tag_config("error", foreground=COLORS['danger'])
        self.log_area.config(state=tk.DISABLED)

        self.toggle_ico_settings()  # 初始隐藏ICO设置

    def toggle_ico_settings(self, event=None):
        """切换ICO尺寸设置的可见性"""
        if self.format_var.get() == 'ICO':
            self.ico_frame.pack(side='left', padx=5)
        else:
            self.ico_frame.pack_forget()

    def select_files(self):
        """选择单个或多个文件"""
        files = filedialog.askopenfilenames(
            title="选择图片文件",
            filetypes=[("图片文件", "*.png *.jpg *.jpeg *.bmp *.webp *.ico")]
        )
        if files:
            self.input_files = list(files)
            self.log_area.config(state=tk.NORMAL)
            self.log_area.insert(tk.END, f"已选择 {len(files)} 个文件\n", "success")
            self.log_area.config(state=tk.DISABLED)

    def select_folder(self):
        """递归选择文件夹中的图片文件"""
        folder = filedialog.askdirectory(title="选择图片文件夹")
        if folder:
            self.input_files = []
            supported_ext = ('png', 'jpg', 'jpeg', 'bmp', 'webp', 'ico')
            for root, _, files in os.walk(folder):
                for f in files:
                    if f.split('.')[-1].lower() in supported_ext:
                        self.input_files.append(os.path.join(root, f))
            self.log_area.config(state=tk.NORMAL)
            self.log_area.insert(tk.END, f"已添加 {len(self.input_files)} 个文件\n", "success")
            self.log_area.config(state=tk.DISABLED)

    def select_output_dir(self):
        """选择输出目录"""
        self.output_dir = filedialog.askdirectory(title="选择输出目录")
        if self.output_dir:
            self.output_label.config(text=self.output_dir)

    def start_conversion(self):
        """启动转换任务"""
        if not self.input_files:
            messagebox.showerror("错误", "请先选择输入文件")
            return
        if not self.output_dir:
            messagebox.showerror("错误", "请选择输出目录")
            return

        self.convert_btn.config(state='disabled', text="⏳ 转换中...")
        self.conversion_running = True
        Thread(target=self.convert_files, daemon=True).start()

    def convert_files(self):
        """执行格式转换核心逻辑"""
        output_format = self.format_var.get().lower()
        ico_size = tuple(map(int, self.size_var.get().split('x'))) if output_format == 'ico' else None

        for input_path in self.input_files:
            if not self.conversion_running:
                break

            try:
                filename = os.path.basename(input_path)
                with Image.open(input_path) as img:
                    # 处理透明通道（JPG/BMP不支持透明）
                    if img.mode in ('RGBA', 'LA') and output_format in ('jpeg', 'bmp'):
                        img = img.convert('RGB')

                    # 调整ICO尺寸
                    if ico_size:
                        img = img.resize(ico_size, Image.Resampling.LANCZOS)

                    # 生成输出路径
                    name = os.path.splitext(filename)[0]
                    output_path = os.path.join(self.output_dir, f"{name}.{output_format}")

                    # 设置保存参数
                    save_args = {'format': output_format}
                    if output_format == 'jpeg':
                        save_args['quality'] = 85  # JPEG质量
                    elif output_format == 'webp':
                        save_args['quality'] = 80  # WEBP质量

                    img.save(output_path, **save_args)
                    self.log_queue.put(("success", f"成功: {filename} → {name}.{output_format}"))
            except Exception as e:
                self.log_queue.put(("error", f"失败: {filename} - {str(e)}"))

        self.log_queue.put(("end", ""))
        self.frame.after(100, lambda: self.convert_btn.config(
            state='normal',
            text="▶ 开始转换"
        ))


# ==================== 超链接转换模块 ====================
class HyperlinkModule(BaseModule):  # 正确类名定义
    def __init__(self, parent):
        self.input_path = tk.StringVar()  # 输入文件路径
        self.output_path = tk.StringVar()  # 输出文件路径
        self.mode = tk.StringVar(value="all")  # 工作表模式
        self.link_mode = tk.StringVar(value="keep")  # 链接显示模式
        self.custom_patterns = {}  # 用户自定义的网盘样式
        self.sheet_names = []  # 工作表列表
        self.config_file = "hyperlink_config.json"  # 配置文件路径

        # 预置常见网盘样式
        self.cloud_storage_patterns = {
            "百度网盘": {"pattern": r"(https?://pan\.baidu\.com/[^\s]+)", "display": "百度网盘资源"},
            "阿里云盘": {"pattern": r"(https?://www\.aliyundrive\.com/[^\s]+)", "display": "阿里云盘资源"},
            "Google Drive": {"pattern": r"(https://drive\.google\.com/[^\s]+)", "display": "Google云端硬盘"},
            "OneDrive": {"pattern": r"(https://\w+\.sharepoint\.com/[^\s]+)", "display": "OneDrive资源"}
        }

        # 加载自定义样式
        self.load_custom_patterns()
        super().__init__(parent)

    def load_custom_patterns(self):
        """从配置文件加载自定义样式"""
        if os.path.exists(self.config_file):
            try:
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    self.custom_patterns = json.load(f)
            except Exception as e:
                messagebox.showerror("错误", f"配置文件加载失败：{str(e)}")

    def save_custom_patterns(self):
        """保存自定义样式到文件"""
        try:
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(self.custom_patterns, f, ensure_ascii=False, indent=2)
        except Exception as e:
            messagebox.showerror("错误", f"配置文件保存失败：{str(e)}")

    def create_widgets(self):
        """构建界面组件（包含新增的样式管理功能）"""
        # ----- 使用说明 -----
        help_text = """使用说明：
1. 选择Excel文件并设置输出路径
2. 添加自定义网盘样式（可选）
3. 选择工作表和处理模式
4. 点击转换按钮执行操作"""
        ttk.Label(self.frame, text=help_text, foreground=COLORS['text']).pack(pady=5, anchor="w")

        # ----- 文件选择区域 -----
        file_frame = ttk.Frame(self.frame)
        # 输入文件
        ttk.Label(file_frame, text="输入文件:").pack(side=tk.LEFT)
        self.input_entry = ttk.Entry(file_frame, textvariable=self.input_path, width=40)
        self.input_entry.pack(side=tk.LEFT, padx=5)
        ttk.Button(file_frame, text="浏览", command=self.select_input).pack(side=tk.LEFT)
        # 输出文件
        ttk.Label(file_frame, text="输出文件:").pack(side=tk.LEFT, padx=10)
        self.output_entry = ttk.Entry(file_frame, textvariable=self.output_path, width=40)
        self.output_entry.pack(side=tk.LEFT, padx=5)
        ttk.Button(file_frame, text="浏览", command=self.select_output).pack(side=tk.LEFT)
        file_frame.pack(pady=10)

        # ----- 自定义网盘设置 -----
        custom_frame = ttk.LabelFrame(self.frame, text="自定义网盘样式（支持正则表达式）")
        # 网盘名称
        ttk.Label(custom_frame, text="网盘名称:").grid(row=0, column=0)
        self.custom_name = ttk.Entry(custom_frame, width=15)
        self.custom_name.grid(row=0, column=1)
        # 正则表达式
        ttk.Label(custom_frame, text="URL正则:").grid(row=0, column=2)
        self.custom_regex = ttk.Entry(custom_frame, width=25)
        self.custom_regex.grid(row=0, column=3)
        # 显示名称
        ttk.Label(custom_frame, text="显示名称:").grid(row=0, column=4)
        self.custom_display = ttk.Entry(custom_frame, width=15)
        self.custom_display.grid(row=0, column=5)
        # 添加按钮
        ttk.Button(custom_frame, text="添加", command=self.add_custom_pattern).grid(row=0, column=6, padx=5)
        custom_frame.pack(pady=5, fill=tk.X)

        # ----- 示例说明 -----
        example_frame = ttk.Frame(self.frame)
        example_text = r"""添加示例：
        网盘名称：我的网盘
        URL正则：mydrive\.com.*
        显示名称：我的私人网盘"""
        ttk.Label(example_frame, text=example_text, foreground=COLORS['secondary']).pack()
        example_frame.pack(pady=5)

        # ----- 处理设置区域 -----
        setting_frame = ttk.Frame(self.frame)
        # 工作表选择
        ttk.Label(setting_frame, text="工作表:").pack(side=tk.LEFT)
        self.sheet_combo = ttk.Combobox(setting_frame, state="readonly", width=15)
        self.sheet_combo.pack(side=tk.LEFT, padx=5)
        # 链接显示模式
        ttk.Label(setting_frame, text="显示模式:").pack(side=tk.LEFT, padx=10)
        ttk.Radiobutton(setting_frame, text="保持链接", variable=self.link_mode, value="keep").pack(side=tk.LEFT)
        ttk.Radiobutton(setting_frame, text="显示为xx网盘", variable=self.link_mode, value="display").pack(side=tk.LEFT)
        setting_frame.pack(pady=10)

        # ----- 操作按钮 -----
        btn_frame = ttk.Frame(self.frame)
        ttk.Button(btn_frame, text="超链接转文本", command=self.convert_to_text, style='Primary.TButton').pack(
            side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="文本转超链接", command=self.convert_to_hyperlink, style='Primary.TButton').pack(
            side=tk.LEFT, padx=5)
        btn_frame.pack(pady=10)

        # ========== 新增：样式管理区域 ==========
        management_frame = ttk.LabelFrame(self.frame, text="已配置样式管理")
        management_frame.pack(pady=10, fill=tk.BOTH, expand=True, padx=10)

        # Treeview组件（显示所有样式）
        self.style_tree = ttk.Treeview(
            management_frame,
            columns=("name", "pattern", "display"),
            show="headings",
            height=5
        )
        self.style_tree.heading("name", text="样式名称")
        self.style_tree.heading("pattern", text="正则表达式")
        self.style_tree.heading("display", text="显示名称")
        self.style_tree.column("name", width=150)
        self.style_tree.column("pattern", width=300)
        self.style_tree.column("display", width=150)
        self.style_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # 右侧操作按钮
        btn_frame = ttk.Frame(management_frame)
        ttk.Button(btn_frame, text="刷新", command=self.refresh_style_list).pack(pady=2, fill=tk.X)
        ttk.Button(btn_frame, text="编辑", command=self.edit_style).pack(pady=2, fill=tk.X)
        ttk.Button(btn_frame, text="删除", command=self.delete_style).pack(pady=2, fill=tk.X)
        btn_frame.pack(side=tk.RIGHT, padx=5)

        # 初始化样式列表
        self.refresh_style_list()

        # ----- 日志区域 -----
        log_frame = ttk.LabelFrame(self.frame, text="操作日志")
        self.log_area = scrolledtext.ScrolledText(log_frame, wrap=tk.WORD, height=15)
        self.log_area.pack(fill=tk.BOTH, expand=True)
        log_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # 配置日志颜色
        self.log_area.tag_config("success", foreground=COLORS['success'])
        self.log_area.tag_config("error", foreground=COLORS['danger'])

    def refresh_style_list(self):
        """刷新样式列表（区分预置和自定义）"""
        self.style_tree.delete(*self.style_tree.get_children())
        # 添加预置样式（灰色显示）
        for name, pattern in self.cloud_storage_patterns.items():
            self.style_tree.insert("", "end", values=(
                f"[预置] {name}",
                pattern["pattern"],
                pattern["display"]
            ), tags=("preset",))
        # 添加自定义样式
        for name, pattern in self.custom_patterns.items():
            self.style_tree.insert("", "end", values=(
                name,
                pattern["pattern"],
                pattern["display"]
            ))
        self.style_tree.tag_configure("preset", foreground="gray")

    def edit_style(self):
        """编辑选中样式"""
        selected = self.style_tree.selection()
        if not selected:
            return

        item = self.style_tree.item(selected[0])
        values = item["values"]

        # 预置样式不可编辑
        if "[预置]" in values[0]:
            messagebox.showinfo("提示", "预置样式不可编辑")
            return

        # 创建编辑窗口
        edit_win = tk.Toplevel()
        edit_win.title("编辑样式")

        # 名称输入框
        ttk.Label(edit_win, text="样式名称:").grid(row=0, column=0, padx=5, pady=5)
        name_entry = ttk.Entry(edit_win)
        name_entry.insert(0, values[0])
        name_entry.grid(row=0, column=1, padx=5, pady=5)

        # 正则输入框
        ttk.Label(edit_win, text="正则表达式:").grid(row=1, column=0, padx=5, pady=5)
        pattern_entry = ttk.Entry(edit_win)
        pattern_entry.insert(0, values[1])
        pattern_entry.grid(row=1, column=1, padx=5, pady=5)

        # 显示名称输入框
        ttk.Label(edit_win, text="显示名称:").grid(row=2, column=0, padx=5, pady=5)
        display_entry = ttk.Entry(edit_win)
        display_entry.insert(0, values[2])
        display_entry.grid(row=2, column=1, padx=5, pady=5)

        def save_changes():
            """保存修改到配置文件"""
            new_name = name_entry.get()
            new_pattern = pattern_entry.get()
            new_display = display_entry.get()

            try:
                re.compile(new_pattern)  # 验证正则表达式
                # 删除旧条目并添加新条目
                del self.custom_patterns[values[0]]
                self.custom_patterns[new_name] = {
                    "pattern": new_pattern,
                    "display": new_display
                }
                self.save_custom_patterns()
                self.refresh_style_list()
                edit_win.destroy()
            except Exception as e:
                messagebox.showerror("错误", f"无效的正则表达式：{str(e)}")

        ttk.Button(edit_win, text="保存", command=save_changes).grid(row=3, columnspan=2, pady=10)

    def delete_style(self):
        """删除选中样式"""
        selected = self.style_tree.selection()
        if not selected:
            return

        item = self.style_tree.item(selected[0])
        values = item["values"]

        if "[预置]" in values[0]:
            messagebox.showinfo("提示", "预置样式不可删除")
            return

        if messagebox.askyesno("确认", f"确定删除样式 '{values[0]}' 吗？"):
            del self.custom_patterns[values[0]]
            self.save_custom_patterns()
            self.refresh_style_list()

    def convert_to_text(self):
        """将超链接转换为文本"""
        try:
            wb = load_workbook(self.input_path.get())
            sheets = self.get_selected_sheets()
            processed = 0

            for sheet_name in sheets:
                ws = wb[sheet_name]
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.hyperlink:
                            cell.value = cell.hyperlink.target
                            cell.hyperlink = None
                            processed += 1

            wb.save(self.output_path.get())
            self.log_area.insert(tk.END, f"成功转换 {processed} 个超链接\n", "success")
        except Exception as e:
            self.log_area.insert(tk.END, f"错误：{str(e)}\n", "error")
            
    def convert_to_hyperlink(self):
        """将文本转换为超链接（支持正则匹配和保持原链接）"""
        try:
            wb = load_workbook(self.input_path.get())
            sheets = self.get_selected_sheets()
            processed = 0

            for sheet_name in sheets:
                ws = wb[sheet_name]
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            url = None
                            display_text = cell.value

                            # 合并所有匹配规则
                            all_patterns = {**self.cloud_storage_patterns, **self.custom_patterns}
                            for name, pattern in all_patterns.items():
                                match = re.search(pattern["pattern"], cell.value)
                                if match:
                                    url = match.group(0)  # 提取匹配到的URL
                                    display_text = pattern["display"] if self.link_mode.get() == "display" else url
                                    break

                            # 保持链接模式：直接使用原文本作为URL
                            if not url and self.link_mode.get() == "keep":
                                url = cell.value

                            if url:
                                cell.hyperlink = url
                                cell.value = display_text
                                cell.font = Font(underline="single", color="0563C1")
                                processed += 1

            wb.save(self.output_path.get())
            self.log_area.insert(tk.END, f"成功转换 {processed} 个链接\n", "success")
        except Exception as e:
            self.log_area.insert(tk.END, f"错误：{str(e)}\n", "error")

    def select_input(self):
        """选择输入文件并加载工作表"""
        file_path = filedialog.askopenfilename(filetypes=[("Excel文件", "*.xlsx")])
        if file_path:
            self.input_path.set(file_path)
            self.output_path.set(file_path.replace(".xlsx", "_转换版.xlsx"))
            self.load_sheets(file_path)

    def select_output(self):
        """选择输出文件路径"""
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx")]
        )
        if file_path:
            self.output_path.set(file_path)

    def load_sheets(self, file_path):
        """加载工作表列表"""
        try:
            wb = load_workbook(file_path, read_only=True)
            self.sheet_names = wb.sheetnames
            self.sheet_combo['values'] = ["全部工作表"] + self.sheet_names
            self.sheet_combo.current(0)
            wb.close()
        except Exception as e:
            messagebox.showerror("错误", f"加载工作表失败：{str(e)}")

    def add_custom_pattern(self):
        """添加自定义网盘样式"""
        name = self.custom_name.get()
        regex = self.custom_regex.get()
        display = self.custom_display.get()

        if not all([name, regex, display]):
            messagebox.showwarning("提示", "请填写所有字段")
            return

        try:
            re.compile(regex)  # 验证正则表达式
            self.custom_patterns[name] = {
                "pattern": regex,
                "display": display
            }
            self.save_custom_patterns()
            self.refresh_style_list()
            self.log_area.insert(tk.END, f"已添加：{name} -> {display}\n", "success")
        except Exception as e:
            messagebox.showerror("错误", f"无效的正则表达式：{str(e)}")

    def get_selected_sheets(self):
        """获取选择的工作表列表"""
        selected = self.sheet_combo.get()
        return self.sheet_names if selected == "全部工作表" else [selected]


# ==================== 程序入口 ====================
if __name__ == "__main__":
    root = tk.Tk()
    app = MainApplication(root)
    root.mainloop()